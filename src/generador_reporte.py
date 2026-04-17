"""
Modulo: generador_reporte.py
Descripcion: Genera el archivo Excel de salida con formato de dashboard.
"""

import pandas as pd
import os
import shutil
from datetime import datetime
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from loguru import logger

COLOR_VERDE    = "27AE60"
COLOR_AMARILLO = "F39C12"
COLOR_ROJO     = "E74C3C"
COLOR_AZUL     = "1A5276"
COLOR_GRIS     = "BDC3C7"
COLOR_BLANCO   = "FFFFFF"

def _aplicar_estilo_header(ws, fila: int, color_hex: str = COLOR_AZUL):
    fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
    font = Font(bold=True, color=COLOR_BLANCO, size=11)
    for cell in ws[fila]:
        if cell.value is not None:
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(horizontal="center", vertical="center")

def _autoajustar_columnas(ws):
    for col in ws.columns:
        max_len = max((len(str(cell.value)) for cell in col if cell.value), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 50)

def _color_semaforo(valor: str) -> str:
    if "Optimo"    in str(valor): return COLOR_VERDE
    if "En Riesgo" in str(valor): return COLOR_AMARILLO
    if "Critico"   in str(valor): return COLOR_ROJO
    return COLOR_GRIS

def generar_reporte(df_detalle: pd.DataFrame, df_resumen: pd.DataFrame, config: dict, metadata: dict) -> str:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    nombre_archivo = f"output_reporte_{timestamp}.xlsx"
    ruta_output = os.path.join(config["rutas"]["output"], nombre_archivo)
    os.makedirs(config["rutas"]["output"], exist_ok=True)
    logger.info(f"Generando reporte: {nombre_archivo}")

    with pd.ExcelWriter(ruta_output, engine="openpyxl") as writer:
        df_resumen.to_excel(writer, sheet_name="Dashboard KPI", index=False)
        ws_dashboard = writer.sheets["Dashboard KPI"]
        _aplicar_estilo_header(ws_dashboard, 1, COLOR_AZUL)
        _autoajustar_columnas(ws_dashboard)

        df_detalle.to_excel(writer, sheet_name="Detalle", index=False)
        ws_detalle = writer.sheets["Detalle"]
        _aplicar_estilo_header(ws_detalle, 1, COLOR_AZUL)
        col_semaforo = df_detalle.columns.get_loc("Semaforo") + 1
        for row_idx, valor in enumerate(df_detalle["Semaforo"], start=2):
            color = _color_semaforo(valor)
            fill  = PatternFill(start_color=color, end_color=color, fill_type="solid")
            ws_detalle.cell(row=row_idx, column=col_semaforo).fill = fill
            ws_detalle.cell(row=row_idx, column=col_semaforo).font = Font(bold=True, color=COLOR_BLANCO)
        _autoajustar_columnas(ws_detalle)

        df_meta = pd.DataFrame(list(metadata.items()), columns=["Parametro", "Valor"])
        df_meta.to_excel(writer, sheet_name="Metadata", index=False)
        ws_meta = writer.sheets["Metadata"]
        _aplicar_estilo_header(ws_meta, 1, COLOR_GRIS)
        _autoajustar_columnas(ws_meta)

    logger.success(f"Reporte generado exitosamente: {ruta_output}")
    return ruta_output

def archivar_fuentes(config: dict) -> None:
    fecha_hoy = datetime.now().strftime("%Y-%m-%d")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    ruta_archive_fecha = os.path.join(config["rutas"]["archive"], fecha_hoy)
    os.makedirs(ruta_archive_fecha, exist_ok=True)
    for archivo_cfg in config["archivos_esperados"]:
        nombre = archivo_cfg["nombre"]
        origen = os.path.join(config["rutas"]["input"], nombre)
        if os.path.exists(origen):
            nombre_base, ext = os.path.splitext(nombre)
            destino = os.path.join(ruta_archive_fecha, f"{nombre_base}_{timestamp}{ext}")
            shutil.move(origen, destino)
            logger.info(f"Archivado: {nombre} -> {destino}")