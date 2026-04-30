"""
Modulo: pipeline.py
Descripcion: Consolidacion de archivos Excel en plantilla de calculo de horas.
"""

import os
import shutil
import unicodedata
from datetime import date, datetime
from pathlib import Path

import pandas as pd
from loguru import logger
from openpyxl import load_workbook
from openpyxl.formula.translate import Translator
from openpyxl.utils import get_column_letter


def _normalizar_texto(texto: str) -> str:
    if texto is None:
        return ""
    texto = str(texto).strip().lower()
    texto = "".join(c for c in unicodedata.normalize("NFD", texto) if unicodedata.category(c) != "Mn")
    return "".join(c for c in texto if c.isalnum())


def _listar_excels(ruta: str) -> list[str]:
    if not os.path.exists(ruta):
        return []
    return [
        f
        for f in os.listdir(ruta)
        if not f.startswith("~$") and f.lower().endswith((".xlsx", ".xlsm", ".xls"))
    ]


def _buscar_archivo_por_keyword(ruta: str, keyword: str) -> str:
    keyword_norm = _normalizar_texto(keyword)
    candidatos = _listar_excels(ruta)
    for nombre in candidatos:
        if keyword_norm in _normalizar_texto(nombre):
            return os.path.join(ruta, nombre)
    raise FileNotFoundError(f"No se encontro archivo con keyword '{keyword}' en {ruta}")


def _buscar_plantilla(ruta_plantillas: str, keyword: str) -> str:
    keyword_norm = _normalizar_texto(keyword)
    candidatos = _listar_excels(ruta_plantillas)
    for nombre in candidatos:
        if keyword_norm in _normalizar_texto(nombre):
            return os.path.join(ruta_plantillas, nombre)
    raise FileNotFoundError(f"No se encontro plantilla con keyword '{keyword}' en {ruta_plantillas}")


def _extraer_columnas_por_indice(ruta_archivo: str, indices_1_based: list[int]) -> pd.DataFrame:
    df = pd.read_excel(ruta_archivo, engine="openpyxl")
    max_cols = df.shape[1]
    for idx in indices_1_based:
        if idx < 1 or idx > max_cols:
            raise ValueError(
                f"El archivo {os.path.basename(ruta_archivo)} no tiene la columna {idx}. Tiene {max_cols} columnas."
            )

    indices_0_based = [i - 1 for i in indices_1_based]
    subset = df.iloc[:, indices_0_based].copy()
    subset.columns = ["col_1", "col_2", "col_3"]
    subset = subset.dropna(how="all")
    return subset


def _resolver_columna_por_header(df: pd.DataFrame, headers_objetivo: list[str]) -> str | None:
    columnas_norm = {_normalizar_texto(col): col for col in df.columns}
    for objetivo in headers_objetivo:
        key = _normalizar_texto(objetivo)
        if key in columnas_norm:
            return columnas_norm[key]
    return None


def _extraer_participantes(ruta_archivo: str) -> pd.DataFrame:
    # La hoja de participantes tiene encabezados en fila 8 y datos desde fila 9.
    df = pd.read_excel(ruta_archivo, engine="openpyxl", header=7)

    col_login = _resolver_columna_por_header(df, ["login"])
    col_training = _resolver_columna_por_header(df, ["training name"])
    col_end_date = _resolver_columna_por_header(df, ["end date of the date", "end date", "end date of data"])
    col_hours = _resolver_columna_por_header(df, ["number of training hours"])

    # Fallback por indice si el encabezado vario en el archivo fuente.
    if col_login is None and df.shape[1] >= 5:
        col_login = df.columns[4]
    if col_training is None and df.shape[1] >= 35:
        col_training = df.columns[34]
    if col_end_date is None and df.shape[1] >= 37:
        col_end_date = df.columns[36]
    if col_hours is None and df.shape[1] >= 41:
        col_hours = df.columns[40]

    if not all([col_login, col_training, col_end_date, col_hours]):
        raise ValueError(
            "No se pudieron resolver columnas requeridas en participantes: "
            f"login={col_login}, training={col_training}, end_date={col_end_date}, hours={col_hours}"
        )

    subset = df[[col_login, col_training, col_end_date, col_hours]].copy()
    subset.columns = ["col_1", "col_2", "col_3", "col_4_hours"]

    # Normaliza AK (End date...) a fecha sin hora para evitar conversion manual en Excel.
    fechas = pd.to_datetime(subset["col_3"], errors="coerce")
    subset["col_3"] = fechas.dt.date.where(fechas.notna(), subset["col_3"])

    subset = subset.dropna(how="all")
    return subset


def _buscar_columna_por_header(ws, header_objetivo: str) -> int | None:
    objetivo = _normalizar_texto(header_objetivo)
    for col in range(1, ws.max_column + 1):
        valor = ws.cell(row=1, column=col).value
        if _normalizar_texto(valor) == objetivo:
            return col
    return None


def _arrastrar_formulas(ws, formula_cols: list[int], fila_base: int, fila_fin: int) -> None:
    if fila_fin <= fila_base:
        return

    for col in formula_cols:
        base_formula = ws.cell(row=fila_base, column=col).value
        if not isinstance(base_formula, str) or not base_formula.startswith("="):
            continue
        origen = f"{get_column_letter(col)}{fila_base}"
        for row in range(fila_base + 1, fila_fin + 1):
            destino = f"{get_column_letter(col)}{row}"
            ws.cell(row=row, column=col).value = Translator(base_formula, origin=origen).translate_formula(destino)


def _archivar_archivos(rutas_archivos: list[str], ruta_archive: str) -> None:
    fecha_hoy = datetime.now().strftime("%Y-%m-%d")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    destino_base = os.path.join(ruta_archive, fecha_hoy)
    os.makedirs(destino_base, exist_ok=True)

    for ruta in rutas_archivos:
        if not os.path.exists(ruta):
            continue
        nombre = Path(ruta).name
        stem = Path(ruta).stem
        sufijo = Path(ruta).suffix
        destino = os.path.join(destino_base, f"{stem}_{timestamp}{sufijo}")
        shutil.move(ruta, destino)
        logger.info(f"Archivado: {nombre} -> {destino}")


def _ultima_fila_ocupada(ws, fila_inicio: int, col_inicio: int, col_fin: int) -> int:
    for fila in range(ws.max_row, fila_inicio - 1, -1):
        if any(ws.cell(row=fila, column=col).value not in (None, "") for col in range(col_inicio, col_fin + 1)):
            return fila
    return fila_inicio - 1


def _generar_reporte_secundario(
    ws_origen,
    ruta_destino_base: str,
    ruta_output_secundario: str,
    hoja_destino: str,
    fila_fin: int,
) -> str:
    # Extrae B:P (cols 2..16) desde fila 3 hasta fila_fin usando el ws ya en memoria.
    datos = []
    for fila in range(3, fila_fin + 1):
        valores = [ws_origen.cell(row=fila, column=col).value for col in range(2, 17)]
        if any(valor not in (None, "") for valor in valores):
            datos.append(valores)

    if not datos:
        raise ValueError("No hay filas con contenido para transferir al proceso secundario.")

    wb_destino = load_workbook(ruta_destino_base)
    if hoja_destino not in wb_destino.sheetnames:
        raise ValueError(f"No existe la hoja '{hoja_destino}' en {os.path.basename(ruta_destino_base)}")
    ws_destino = wb_destino[hoja_destino]

    fila_inicio_pegado = _ultima_fila_ocupada(ws_destino, fila_inicio=2, col_inicio=1, col_fin=15) + 1
    if fila_inicio_pegado < 2:
        fila_inicio_pegado = 2

    for offset, valores in enumerate(datos):
        fila_objetivo = fila_inicio_pegado + offset
        for col_offset, valor in enumerate(valores):
            ws_destino.cell(row=fila_objetivo, column=1 + col_offset).value = valor

    os.makedirs(ruta_output_secundario, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    nombre_base = Path(ruta_destino_base).stem
    ruta_salida_secundaria = os.path.join(ruta_output_secundario, f"{nombre_base}_actualizado_{timestamp}.xlsx")
    wb_destino.save(ruta_salida_secundaria)
    return ruta_salida_secundaria


def ejecutar_pipeline(config: dict) -> None:
    inicio = datetime.now()
    logger.info("=" * 60)
    logger.info("INICIANDO CONSOLIDACION DE EXCEL")
    logger.info("=" * 60)

    ruta_input = config["rutas"]["input"]
    ruta_plantillas = config["rutas"].get("plantillas", "./data/plantillas")
    ruta_output = config["rutas"]["output"]
    ruta_input_secundario = config["rutas"].get("input_secundario", "./data/input_secundario")
    ruta_output_secundario = config["rutas"].get("output_secundario", "./data/output_secundario")
    ruta_archive = config["rutas"]["archive"]

    cfg_excel = config.get("consolidacion_excel", {})
    kw_elearning = cfg_excel.get("keyword_elearning", "elearning")
    kw_files_jmc = cfg_excel.get("keyword_files_jmc", "filesjmc")
    kw_participantes = cfg_excel.get("keyword_participantes", "participantes")
    kw_plantilla = cfg_excel.get("keyword_plantilla", "calculo horas 2026")
    hoja_destino = cfg_excel.get("hoja_destino", "CONSOLIDADO")

    cfg_secundario = config.get("proceso_secundario", {})
    sec_habilitado = cfg_secundario.get("habilitado", True)
    kw_secundario = cfg_secundario.get("keyword_archivo_destino", "horas entrenamiento 2026")
    hoja_secundaria = cfg_secundario.get("hoja_destino", "Reporte horas")

    os.makedirs(ruta_output, exist_ok=True)

    try:
        logger.info("PASO 1: Identificacion de archivos por keyword...")
        ruta_elearning = _buscar_archivo_por_keyword(ruta_input, kw_elearning)
        ruta_files_jmc = _buscar_archivo_por_keyword(ruta_input, kw_files_jmc)
        ruta_participantes = _buscar_archivo_por_keyword(ruta_input, kw_participantes)
        ruta_plantilla = _buscar_plantilla(ruta_plantillas, kw_plantilla)

        logger.info(f"Archivo A (elearning): {os.path.basename(ruta_elearning)}")
        logger.info(f"Archivo B (filesJMC): {os.path.basename(ruta_files_jmc)}")
        logger.info(f"Archivo C (participantes): {os.path.basename(ruta_participantes)}")
        logger.info(f"Plantilla detectada: {os.path.basename(ruta_plantilla)}")

        logger.info("PASO 2: Extraccion de columnas y consolidacion continua...")
        df_elearning = _extraer_columnas_por_indice(ruta_elearning, [1, 2, 9])
        df_elearning["col_4_hours"] = pd.NA
        df_files_jmc = _extraer_columnas_por_indice(ruta_files_jmc, [1, 4, 13])
        df_files_jmc["col_4_hours"] = pd.NA
        df_participantes = _extraer_participantes(ruta_participantes)
        df_consolidado = pd.concat([df_elearning, df_files_jmc, df_participantes], ignore_index=True)

        if df_consolidado.empty:
            raise ValueError("No se encontraron filas para consolidar en la plantilla.")

        logger.info(
            "PASO 3: Pegado de valores en plantilla "
            "(sin encabezados, desde fila 3 en columnas B, E y H; participantes AO -> J)..."
        )
        wb = load_workbook(ruta_plantilla)
        ws = wb[hoja_destino] if hoja_destino in wb.sheetnames else wb[wb.sheetnames[0]]

        header_inicio = ws.cell(row=2, column=2).value
        header_objeto = ws.cell(row=2, column=5).value
        header_fecha = ws.cell(row=2, column=8).value

        logger.info(
            "Encabezados objetivo detectados en fila 2: "
            f"B2='{header_inicio}', E2='{header_objeto}', H2='{header_fecha}'"
        )

        col_inicio, col_objeto, col_fecha = 2, 5, 8
        fila_inicio = 3
        fila_fin = fila_inicio + len(df_consolidado) - 1
        fila_fin_files_jmc = fila_inicio + len(df_elearning) + len(df_files_jmc) - 1

        for i, row in enumerate(df_consolidado.itertuples(index=False), start=fila_inicio):
            ws.cell(row=i, column=col_inicio).value = row.col_1
            celda_e = ws.cell(row=i, column=col_objeto)
            celda_e.value = row.col_3
            if isinstance(row.col_3, date):
                celda_e.number_format = "dd/mm/yyyy"
            ws.cell(row=i, column=col_fecha).value = row.col_2
            if pd.notna(row.col_4_hours):
                ws.cell(row=i, column=10).value = row.col_4_hours

        logger.info("PASO 4: Arrastre de formulas hasta la ultima fila con datos nuevos...")
        formula_cols = [
            col
            for col in range(1, ws.max_column + 1)
            if isinstance(ws.cell(row=3, column=col).value, str) and ws.cell(row=3, column=col).value.startswith("=")
        ]

        # En columna J (10), mantener formulas solo para filas de elearning+filesJMC.
        # En filas de participantes se conservan los valores crudos de AO sin sobreescribir.
        if 10 in formula_cols:
            _arrastrar_formulas(ws, [10], fila_base=3, fila_fin=fila_fin_files_jmc)
            formula_cols = [col for col in formula_cols if col != 10]

        _arrastrar_formulas(ws, formula_cols, fila_base=3, fila_fin=fila_fin)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        nombre_base = Path(ruta_plantilla).stem
        ruta_salida = os.path.join(ruta_output, f"{nombre_base}_consolidado_{timestamp}.xlsx")
        wb.save(ruta_salida)

        logger.info("PASO 5: Archivando archivos de entrada procesados...")
        _archivar_archivos([ruta_elearning, ruta_files_jmc, ruta_participantes], ruta_archive)

        if sec_habilitado:
            logger.info(
                "PASO 6: Proceso secundario (copiar B:P desde consolidado a hoja Reporte horas en A:O, desde fila 2)..."
            )
            ruta_base_secundaria = _buscar_archivo_por_keyword(ruta_input_secundario, kw_secundario)
            ruta_salida_secundaria = _generar_reporte_secundario(
                ws_origen=ws,
                ruta_destino_base=ruta_base_secundaria,
                ruta_output_secundario=ruta_output_secundario,
                hoja_destino=hoja_secundaria,
                fila_fin=fila_fin,
            )
            logger.success(f"Archivo secundario generado: {ruta_salida_secundaria}")

        fin = datetime.now()
        logger.info("=" * 60)
        logger.success(f"CONSOLIDACION COMPLETADA en {round((fin - inicio).total_seconds(), 2)}s")
        logger.success(f"Archivo generado: {ruta_salida}")
        logger.info("=" * 60)

    except Exception as e:
        logger.error(f"ERROR en la consolidacion: {e}")
        raise
